import os, build_workbook
from tkinter import filedialog
from tkinter import *
from openpyxl import load_workbook

def main():
    # Default Arguments
    argument_json = ""
    argument_xlsx = ""
    argument_output = ""
    argument_substring_filename = "OUTPUT.xlsx"
    argument_row = 0
    argument_col = 0

    # Root Init
    root = Tk()
    root.title("JSON to XLSX Utility (WIP)")
    root.geometry('800x600')

    # Label Vars
    var_output = StringVar()
    var_output.set("Select Output Directory")
    var_template = StringVar()
    var_template.set("Select Template File")
    var_contents = StringVar()
    var_contents.set("Contents will appear here")
    var_status = StringVar()
    var_status.set("Spreadsheet Generator")
    var_echo = StringVar()
    var_echo.set("")

    # Labels
    Label(textvariable=var_template).grid(row=1, column=1)
    Label(textvariable=var_output).grid(row=2, column=1)
    Label(textvariable=var_contents).grid(row=4, column=0)
    Label(textvariable=var_status).grid(row=0, column=0)
    Label(textvariable=var_echo).grid(row=4, column=1)

    def sel_template():
        global argument_xlsx
        argument_xlsx = filedialog.askopenfilename(initialdir = "/",title = "Select template",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
        var_template.set("Template File: " + argument_xlsx)
        print(argument_xlsx)

    def sel_json():
        global argument_json
        argument_json = filedialog.askdirectory(initialdir="/", title="Select directory")
        var_output.set("JSON Directory: " + argument_json)
        temp="Contents:\n"
        for file in os.listdir(argument_json):
            if file.endswith(".json"):
                temp += file
                temp += "\n"
        var_contents.set(temp)

    def gen_sheet():
        global argument_json, argument_xlsx, argument_output, argument_row, argument_col, var_status
        print("TEST:" + argument_xlsx)
        wb = load_workbook(filename=argument_xlsx)
        ws = wb.active

        # Some variables are defined programmatically for now

        argument_output = argument_json + "/" + argument_substring_filename
        if(isinstance(ws['A1'].value, int)):
            argument_row = ws['A1'].value
        if (isinstance(ws['A2'].value, int)):
            argument_col = ws['A2'].value
        var_echo.set(build_workbook.main(argument_json, argument_xlsx, argument_output, argument_row, argument_col))

    button_template = Button(root, width=40, text="Select XLSX Output Template", command=sel_template).grid(row=1, column=0)
    button_output= Button(root, width=40, text="Select Folder Containing .JSON files", command=sel_json).grid(row=2, column=0)
    button_gen = Button(root, width=40, text="Generate Spreadsheet", command=gen_sheet).grid(row=3, column=0)

    root.mainloop()

if __name__ == "__main__":
    main()