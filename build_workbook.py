import json, os, argparse
from openpyxl import load_workbook

# This is a little command line tool that allows you to turn JSON files
# into a nice formatted spreadsheet.
# It is a bit messy and not flexable, however it works for the time
# being. Feel free to modify it as needed. There is a GUI component as well.
# ~ Jon, 5/6/2019

def main(_json, _xlsx, _output, _row, _col):
    wb = load_workbook(filename=_xlsx)
    ws1 = wb.active
    ws1.title = "Data"
    active_row = _row
    active_col = _col

    # -- VARS
    # This is very messy but it gets the job done for now
    # They are hardcoded to look for specific variables, change as needed

    SYSTEM = "001#Machine"
    CHIP = "003#CPU"
    GRAPHICS = "004#Graphics"
    AUDIO = "005#Audio"
    NETWORK = "006#Network"
    STORAGE = "007#Drives"
    MEM = "011#Info"
    CUST = "CUSTOM_INFO"

    SYSTEM_SYSTEM = [SYSTEM, 0, "001#System"]
    SYSTEM_PRODUCTNAME = [SYSTEM, 0, "002#product"]
    SYSTEM_SERIALNUMBER = [SYSTEM, 0, "004#serial"]
    CHIP_NAME = [CHIP, 0, "001#model"]
    CHIP_BITS = [CHIP, 0, "002#bits"]
    CHIP_TYPE = [CHIP, 0, "003#type"]
    CHIP_ARCH = [CHIP, 0, "004#arch"]
    CHIP_CLOCKSPEED = [CHIP, 2, "010#min/max"]
    GRAPHICS_NAME = [GRAPHICS, 0, "001#Device"]
    GRAPHICS_RES = [GRAPHICS, 1, "004#resolution"]
    AUDIO_NAME = [AUDIO, 0, "001#Device"]
    NETWORK_NAME = [NETWORK, 0, "001#Device"]
    NETWORK_SPEED = [NETWORK, 1, "002#speed"]
    NETWORK_WIRELESS = [NETWORK, 2, "001#Device"]
    STORAGE_NAME = [STORAGE, 1, "003#model"]
    STORAGE_SIZE = [STORAGE, 1, "004#size"]
    MEM_SIZE = [MEM, 0, "002#Memory"]
    NOTES = [CUST, 0, "CUST_NOTES"]

    # Format that will be added to the sheet (Starting at _row, _col)
    OUTPUT = [SYSTEM_SYSTEM, SYSTEM_PRODUCTNAME, SYSTEM_SERIALNUMBER, CHIP_NAME, CHIP_BITS, CHIP_TYPE, CHIP_ARCH,
              CHIP_CLOCKSPEED,
              GRAPHICS_NAME, GRAPHICS_RES, AUDIO_NAME, NETWORK_NAME, NETWORK_SPEED, NETWORK_WIRELESS, STORAGE_NAME,
              STORAGE_SIZE, MEM_SIZE, NOTES]

    def write_data(path):
        with open(path) as data_file:
            data = json.load(data_file)
            col = active_col
            for _format in OUTPUT:
                ws1.cell(column=col, row=active_row, value=extract_data(data, _format))
                col += 1

    def extract_data(data, _format):
        try:
            return data[_format[0]][_format[1]][_format[2]]
        except:
            return 'Error'

    output = "Process Complete!\nFiles Marked with a '+' have been added to the sheet.\nFiles Marked with a '-' have been ignored.\nList of Files/Folders in directory:\n"
    for file in os.listdir(_json):
        if file.endswith(".json"):
            output += "+ " + file + "\n"
            write_data(os.path.join(_json, file))
            active_row += 1
        else:
            output += "- " + file + "\n"
    wb.save(filename=_output)
    print(output)
    return output

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Tool for writing json data to spreadsheet')
    parser.add_argument("json", help="Directory where json files are located")
    parser.add_argument("xlsx", help="Path to template xlsx file")
    parser.add_argument("output", help="Output directory and file name")
    parser.add_argument("row", help="Starting row", type=int)
    parser.add_argument("col", help="Starting col", type=int)
    args = parser.parse_args()
    main(args.json, args.xlsx, args.output, args.row, args.col)
